"""Export service for Excel and CSV generation"""
import csv
import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

from app.models.item import Item
from app.models.batch import Batch
from app.models.movement import Movement


class ExportService:
    """Service for exporting data to Excel and CSV"""
    
    @staticmethod
    def create_excel_response(workbook: Workbook, filename: str) -> StreamingResponse:
        """Create StreamingResponse from Excel workbook"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    
    @staticmethod
    def create_csv_response(data: str, filename: str) -> StreamingResponse:
        """Create StreamingResponse from CSV data"""
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            io.StringIO(data),
            media_type='text/csv',
            headers=headers
        )
    
    @staticmethod
    def export_items_excel(items: List[Item]) -> StreamingResponse:
        """Export items to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        
        # Header style
        header_fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "SKU", "Name", "Description", "Supplier", 
            "Unit of Measure", "Cost Price", "Currency",
            "Reorder Point", "Min Stock", "Max Stock",
            "Created At", "Updated At"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.sku)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.description)
            ws.cell(row=row, column=4, value=item.supplier)
            ws.cell(row=row, column=5, value=item.unit_of_measure)
            ws.cell(row=row, column=6, value=float(item.cost_price))
            ws.cell(row=row, column=7, value=item.currency)
            ws.cell(row=row, column=8, value=item.reorder_point)
            ws.cell(row=row, column=9, value=item.min_stock)
            ws.cell(row=row, column=10, value=item.max_stock)
            ws.cell(row=row, column=11, value=item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "")
            ws.cell(row=row, column=12, value=item.updated_at.strftime("%Y-%m-%d %H:%M") if item.updated_at else "")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        filename = f"items_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExportService.create_excel_response(wb, filename)
    
    @staticmethod
    def export_items_csv(items: List[Item]) -> StreamingResponse:
        """Export items to CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "SKU", "Name", "Description", "Supplier",
            "Unit of Measure", "Cost Price", "Currency",
            "Reorder Point", "Min Stock", "Max Stock",
            "Created At", "Updated At"
        ])
        
        # Data rows
        for item in items:
            writer.writerow([
                item.sku,
                item.name,
                item.description or "",
                item.supplier,
                item.unit_of_measure,
                float(item.cost_price),
                item.currency,
                item.reorder_point,
                item.min_stock,
                item.max_stock,
                item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "",
                item.updated_at.strftime("%Y-%m-%d %H:%M") if item.updated_at else ""
            ])
        
        filename = f"items_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return ExportService.create_csv_response(output.getvalue(), filename)
    
    @staticmethod
    def export_batches_excel(batches: List[Batch]) -> StreamingResponse:
        """Export batches to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Batches"
        
        # Header style
        header_fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "Batch Number", "Item Name", "Item SKU",
            "Quantity Received", "Quantity Available", "Status",
            "Intake Date", "Expiration Date", "Days Until Expiry",
            "Location", "Notes", "Created At"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        today = datetime.now().date()
        for row, batch in enumerate(batches, 2):
            days_until_expiry = (batch.expiration_date - today).days if batch.expiration_date else None
            
            ws.cell(row=row, column=1, value=batch.batch_number)
            ws.cell(row=row, column=2, value=batch.item.name if batch.item else "")
            ws.cell(row=row, column=3, value=batch.item.sku if batch.item else "")
            ws.cell(row=row, column=4, value=float(batch.quantity_received))
            ws.cell(row=row, column=5, value=float(batch.quantity_available))
            ws.cell(row=row, column=6, value=batch.status.value)
            ws.cell(row=row, column=7, value=batch.receipt_date.strftime("%Y-%m-%d") if batch.receipt_date else "")
            ws.cell(row=row, column=8, value=batch.expiration_date.strftime("%Y-%m-%d") if batch.expiration_date else "")
            ws.cell(row=row, column=9, value=days_until_expiry)
            ws.cell(row=row, column=10, value=batch.location.location_code if batch.location else "")
            ws.cell(row=row, column=11, value=batch.notes or "")
            ws.cell(row=row, column=12, value=batch.created_at.strftime("%Y-%m-%d %H:%M") if batch.created_at else "")
            
            # Color-code by expiration status
            if days_until_expiry is not None:
                if days_until_expiry < 0:
                    fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif days_until_expiry <= 30:
                    fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                else:
                    fill = None
                
                if fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).fill = fill
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        filename = f"batches_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExportService.create_excel_response(wb, filename)
    
    @staticmethod
    def export_batches_csv(batches: List[Batch]) -> StreamingResponse:
        """Export batches to CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "Batch Number", "Item Name", "Item SKU",
            "Quantity Received", "Quantity Available", "Status",
            "Intake Date", "Expiration Date", "Days Until Expiry",
            "Location", "Notes", "Created At"
        ])
        
        # Data rows
        today = datetime.now().date()
        for batch in batches:
            days_until_expiry = (batch.expiration_date - today).days if batch.expiration_date else None
            
            writer.writerow([
                batch.batch_number,
                batch.item.name if batch.item else "",
                batch.item.sku if batch.item else "",
                float(batch.quantity_received),
                float(batch.quantity_available),
                batch.status.value,
                batch.receipt_date.strftime("%Y-%m-%d") if batch.receipt_date else "",
                batch.expiration_date.strftime("%Y-%m-%d") if batch.expiration_date else "",
                days_until_expiry or "",
                batch.location.location_code if batch.location else "",
                batch.notes or "",
                batch.created_at.strftime("%Y-%m-%d %H:%M") if batch.created_at else ""
            ])
        
        filename = f"batches_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return ExportService.create_csv_response(output.getvalue(), filename)
    
    @staticmethod
    def export_movements_excel(movements: List[Movement]) -> StreamingResponse:
        """Export movements to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Movements"
        
        # Header style
        header_fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "Timestamp", "Batch Number", "Item Name",
            "Movement Type", "Quantity", "Quantity Before",
            "Quantity After", "User", "Reference Number", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row, movement in enumerate(movements, 2):
            ws.cell(row=row, column=1, value=movement.timestamp.strftime("%Y-%m-%d %H:%M") if movement.timestamp else "")
            ws.cell(row=row, column=2, value=movement.batch.batch_number if movement.batch else "")
            ws.cell(row=row, column=3, value=movement.batch.item.name if movement.batch and movement.batch.item else "")
            ws.cell(row=row, column=4, value=movement.movement_type.value)
            ws.cell(row=row, column=5, value=float(movement.quantity))
            ws.cell(row=row, column=6, value=float(movement.quantity_before))
            ws.cell(row=row, column=7, value=float(movement.quantity_after))
            ws.cell(row=row, column=8, value=movement.user.username if movement.user else "")
            ws.cell(row=row, column=9, value=movement.reference_number or "")
            ws.cell(row=row, column=10, value=movement.notes or "")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        filename = f"movements_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExportService.create_excel_response(wb, filename)
    
    @staticmethod
    def export_movements_csv(movements: List[Movement]) -> StreamingResponse:
        """Export movements to CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "Timestamp", "Batch Number", "Item Name",
            "Movement Type", "Quantity", "Quantity Before",
            "Quantity After", "User", "Reference Number", "Notes"
        ])
        
        # Data rows
        for movement in movements:
            writer.writerow([
                movement.timestamp.strftime("%Y-%m-%d %H:%M") if movement.timestamp else "",
                movement.batch.batch_number if movement.batch else "",
                movement.batch.item.name if movement.batch and movement.batch.item else "",
                movement.movement_type.value,
                float(movement.quantity),
                float(movement.quantity_before),
                float(movement.quantity_after),
                movement.user.username if movement.user else "",
                movement.reference_number or "",
                movement.notes or ""
            ])
        
        filename = f"movements_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return ExportService.create_csv_response(output.getvalue(), filename)


export_service = ExportService()
